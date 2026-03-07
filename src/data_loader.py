"""Load and normalize personal finance data from the Excel workbook."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import io
import logging
from time import perf_counter
import re
from pathlib import Path

import openpyxl
from openpyxl.workbook.workbook import Workbook
import pandas as pd

LOGGER = logging.getLogger(__name__)


MONTH_INDEX = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


@dataclass(frozen=True)
class WorkbookData:
    """Container with all normalized DataFrames used by the app."""

    raw_transactions: pd.DataFrame
    general_summary: pd.DataFrame
    expenses_by_category: pd.DataFrame
    budget: pd.DataFrame
    income_by_source: pd.DataFrame
    category_averages: pd.DataFrame


WorkbookSource = str | Path | bytes | Workbook | None


def resolve_workbook_path(workbook_path: str | Path | None = None) -> Path:
    """Resolve the workbook path from explicit or default locations."""
    LOGGER.debug("Resolving workbook path. workbook_path=%s", workbook_path)
    if workbook_path is not None:
        resolved = Path(workbook_path)
        LOGGER.debug("Using explicit workbook path: %s", resolved)
        return resolved

    root = Path(__file__).resolve().parent.parent
    candidates = [
        root / "data" / "Mock Personal finances.xlsx",
        root / "Mock Personal finances.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            LOGGER.debug("Resolved workbook path from defaults: %s", candidate)
            return candidate
    LOGGER.warning("Workbook not found in default locations. Returning expected path: %s", candidates[0])
    return candidates[0]


def _load_workbook(source: WorkbookSource = None) -> Workbook:
    LOGGER.debug("Loading workbook. source_type=%s", type(source).__name__)
    try:
        if isinstance(source, Workbook):
            LOGGER.debug("Workbook object provided directly; reusing open workbook instance.")
            return source
        if isinstance(source, bytes):
            workbook = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
            LOGGER.debug("Workbook loaded from in-memory bytes. byte_size=%s", len(source))
            return workbook
        path = resolve_workbook_path(source)
        workbook = openpyxl.load_workbook(path, data_only=True)
        LOGGER.debug("Workbook loaded from path: %s", path)
        return workbook
    except Exception:
        LOGGER.error("Failed to open workbook from source_type=%s", type(source).__name__, exc_info=True)
        raise


def _to_float(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return default
    return default


def _parse_month_value(value: object) -> pd.Timestamp | None:
    if isinstance(value, pd.Timestamp):
        return value.replace(day=1).normalize()
    if isinstance(value, datetime):
        return pd.Timestamp(value.year, value.month, 1)
    if isinstance(value, date):
        return pd.Timestamp(value.year, value.month, 1)
    if not isinstance(value, str):
        return None

    text = value.strip().lower()
    compact = re.sub(r"[\s_\-]", "", text)

    match = re.match(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)(\d{4})$", compact)
    if match:
        month = MONTH_INDEX[match.group(1)]
        year = int(match.group(2))
        return pd.Timestamp(year, month, 1)

    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return pd.Timestamp(parsed.year, parsed.month, 1)
    except Exception:
        return None
    return None


def _extract_month_columns(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    start_col: int = 3,
) -> list[tuple[int, pd.Timestamp]]:
    months: list[tuple[int, pd.Timestamp]] = []
    for col in range(start_col, worksheet.max_column + 1):
        month_value = _parse_month_value(worksheet.cell(row=header_row, column=col).value)
        if month_value is None:
            continue
        months.append((col, month_value))
    if not months:
        LOGGER.warning(
            "No month columns detected in worksheet=%s header_row=%s start_col=%s",
            worksheet.title,
            header_row,
            start_col,
        )
    else:
        LOGGER.debug("Detected %s month columns in worksheet=%s", len(months), worksheet.title)
    return months


def load_raw_transactions(source: WorkbookSource = None) -> pd.DataFrame:
    """Load transaction-level rows from RAW sheet and normalize values."""
    LOGGER.debug("load_raw_transactions started. source_type=%s", type(source).__name__)
    ws = _load_workbook(source)["RAW"]
    records: list[dict[str, object]] = []

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        raw_date, category, amount, description = row
        if category is None or amount is None:
            continue
        if not isinstance(raw_date, (datetime, date)):
            continue
        if raw_date.year >= 2100:
            continue

        signed_amount = _to_float(amount)
        records.append(
            {
                "date": pd.Timestamp(raw_date),
                "category": str(category),
                "amount": abs(signed_amount),
                "signed_amount": signed_amount,
                "description": description if description is not None else "",
            }
        )

    frame = pd.DataFrame(records)
    if frame.empty:
        LOGGER.warning("RAW sheet produced no transaction rows.")
        return pd.DataFrame(columns=["date", "category", "amount", "signed_amount", "description"])

    frame["year"] = frame["date"].dt.year
    frame["month"] = frame["date"].dt.to_period("M").dt.to_timestamp()
    frame = frame.sort_values("date").reset_index(drop=True)
    LOGGER.debug("load_raw_transactions completed. rows=%s cols=%s", len(frame), len(frame.columns))
    return frame


def load_general_summary(source: WorkbookSource = None) -> pd.DataFrame:
    """Load monthly INCOME/EXPENSES/INVESTMENTS/SAVINGS totals from GENERAL."""
    LOGGER.debug("load_general_summary started. source_type=%s", type(source).__name__)
    ws = _load_workbook(source)["GENERAL"]
    months = _extract_month_columns(ws, header_row=20, start_col=3)
    source_rows = [21, 22, 23, 24]
    records: list[dict[str, object]] = []

    for row_idx in source_rows:
        source = ws.cell(row=row_idx, column=2).value
        status = ws.cell(row=row_idx, column=1).value
        if source is None:
            continue
        for col_idx, month in months:
            amount = _to_float(ws.cell(row=row_idx, column=col_idx).value)
            records.append(
                {
                    "month": month,
                    "source": str(source).upper(),
                    "amount": amount,
                    "status": status,
                }
            )

    frame = pd.DataFrame(records)
    if frame.empty:
        LOGGER.warning("GENERAL summary produced no rows.")
    LOGGER.debug("load_general_summary completed. rows=%s cols=%s", len(frame), len(frame.columns))
    return frame


def load_expenses_by_category(source: WorkbookSource = None) -> pd.DataFrame:
    """Load monthly expense totals by category from EXPENSES historical section."""
    LOGGER.debug("load_expenses_by_category started. source_type=%s", type(source).__name__)
    ws = _load_workbook(source)["EXPENSES"]
    months = _extract_month_columns(ws, header_row=48, start_col=3)

    categories: list[str] = []
    for row_idx in range(22, 35):
        category = ws.cell(row=row_idx, column=2).value
        if category is None:
            continue
        category_str = str(category).strip()
        if category_str.upper() == "TOTAL":
            continue
        categories.append(category_str)
    if not categories:
        LOGGER.warning("No expense categories found in EXPENSES historical section.")

    start_row = 50
    records: list[dict[str, object]] = []
    for offset, category in enumerate(categories):
        row_idx = start_row + offset
        for col_idx, month in months:
            amount = _to_float(ws.cell(row=row_idx, column=col_idx).value)
            records.append(
                {
                    "month": month,
                    "category": category,
                    "amount": amount,
                }
            )

    frame = pd.DataFrame(records)
    if not frame.empty:
        frame = frame.sort_values(["month", "category"]).reset_index(drop=True)
    else:
        LOGGER.warning("EXPENSES by category produced no rows.")
    LOGGER.debug("load_expenses_by_category completed. rows=%s cols=%s", len(frame), len(frame.columns))
    return frame


def load_budget(source: WorkbookSource = None) -> pd.DataFrame:
    """Load projected vs actual budget values from EXPENSES current section."""
    LOGGER.debug("load_budget started. source_type=%s", type(source).__name__)
    ws = _load_workbook(source)["EXPENSES"]
    year = int(_to_float(ws.cell(row=3, column=1).value, default=0)) or datetime.now().year
    month_name = str(ws.cell(row=4, column=1).value or "").strip().lower()[:3]
    month = MONTH_INDEX.get(month_name, 1)
    budget_month = pd.Timestamp(year, month, 1)

    records: list[dict[str, object]] = []
    for row_idx in range(4, 18):
        category = ws.cell(row=row_idx, column=2).value
        if category is None:
            continue
        category_str = str(category).strip()
        records.append(
            {
                "month": budget_month,
                "category": category_str,
                "projected": _to_float(ws.cell(row=row_idx, column=3).value),
                "actual": _to_float(ws.cell(row=row_idx, column=4).value),
                "difference": _to_float(ws.cell(row=row_idx, column=7).value),
                "is_total": category_str.upper() == "TOTAL",
            }
        )

    frame = pd.DataFrame(records)
    if frame.empty:
        LOGGER.warning("Budget section produced no rows.")
    LOGGER.debug("load_budget completed. rows=%s cols=%s", len(frame), len(frame.columns))
    return frame


def load_income_by_source(source: WorkbookSource = None) -> pd.DataFrame:
    """Load monthly income by source from INCOME historical section."""
    LOGGER.debug("load_income_by_source started. source_type=%s", type(source).__name__)
    ws = _load_workbook(source)["INCOME"]
    months = _extract_month_columns(ws, header_row=23, start_col=3)
    source_rows = range(24, 29)
    records: list[dict[str, object]] = []

    for row_idx in source_rows:
        source = ws.cell(row=row_idx, column=2).value
        status = ws.cell(row=row_idx, column=1).value
        if source is None:
            continue
        source_str = str(source).strip()
        for col_idx, month in months:
            amount = _to_float(ws.cell(row=row_idx, column=col_idx).value)
            records.append(
                {
                    "month": month,
                    "source": source_str,
                    "amount": amount,
                    "status": status,
                    "is_total": source_str.upper() == "TOTAL",
                }
            )

    frame = pd.DataFrame(records)
    if frame.empty:
        LOGGER.warning("INCOME by source produced no rows.")
    LOGGER.debug("load_income_by_source completed. rows=%s cols=%s", len(frame), len(frame.columns))
    return frame


def load_category_averages(source: WorkbookSource = None) -> pd.DataFrame:
    """Load yearly average amounts by expense sub-category."""
    LOGGER.debug("load_category_averages started. source_type=%s", type(source).__name__)
    ws = _load_workbook(source)["EXPENSES"]
    years: list[tuple[int, int]] = []
    for col in range(3, ws.max_column + 1):
        value = ws.cell(row=38, column=col).value
        year: int | None = None
        if isinstance(value, (int, float)):
            year = int(value)
        elif isinstance(value, str) and value.strip().isdigit():
            year = int(value.strip())
        if year is not None and 2000 <= year <= 2100:
            years.append((col, year))
    if not years:
        LOGGER.warning("No valid yearly columns found for category averages.")

    records: list[dict[str, object]] = []
    for row_idx in range(40, 45):
        category = ws.cell(row=row_idx, column=2).value
        if not category:
            continue
        for col_idx, year in years:
            records.append(
                {
                    "category": str(category),
                    "year": year,
                    "average_amount": _to_float(ws.cell(row=row_idx, column=col_idx).value),
                }
            )

    frame = pd.DataFrame(records)
    if frame.empty:
        LOGGER.warning("Category averages produced no rows.")
    LOGGER.debug("load_category_averages completed. rows=%s cols=%s", len(frame), len(frame.columns))
    return frame


def load_all_data(source: WorkbookSource = None) -> WorkbookData:
    """Load all DataFrames required by the dashboard."""
    LOGGER.debug("load_all_data started. source_type=%s", type(source).__name__)
    start_time = perf_counter()
    workbook = _load_workbook(source)
    data = WorkbookData(
        raw_transactions=load_raw_transactions(workbook),
        general_summary=load_general_summary(workbook),
        expenses_by_category=load_expenses_by_category(workbook),
        budget=load_budget(workbook),
        income_by_source=load_income_by_source(workbook),
        category_averages=load_category_averages(workbook),
    )
    elapsed_ms = (perf_counter() - start_time) * 1000
    LOGGER.info(
        "Workbook data loaded in %.2fms.",
        elapsed_ms,
    )
    LOGGER.debug(
        (
            "load_all_data details. raw_transactions=%s general_summary=%s "
            "expenses_by_category=%s budget=%s income_by_source=%s category_averages=%s"
        ),
        len(data.raw_transactions),
        len(data.general_summary),
        len(data.expenses_by_category),
        len(data.budget),
        len(data.income_by_source),
        len(data.category_averages),
    )
    return data
